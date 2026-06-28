"""Build DETAILED_TABLE.xlsx (+ .md) for the LATEST live fused virtual-axis backend.

Same layout as the archived reports_archive/per_step_active_state/DETAILED_TABLE.xlsx, but
**Clifft baseline vs live fused-VA only** (no TTN, no near-Clifford block).  All fused numbers
are read straight from the per-step traces emitted by `FusedSingleFrame`
(`fused_va_<circuit>_per_step.csv`), i.e. the dense-free single-frame backend's OWN run.

Three sheets:
  1. Transient vs Resident  -- PEAK active-state, Clifft 2^k vs fused transient / resident.
  2. Active-State (2^x)      -- PEAK and time-integrated SUM, dense-equivalent dimension.
  3. Memory (bytes)          -- PEAK and SUM, dense magic register (complex128 = 16 B/amp).

state size  = active-state dimension 2^b (the dense magic register the fused engine holds).
memory size = 16 * 2^b bytes (complex128).  The O(n^2)-bit CHP tableau is excluded (poly, the
              same basis on which Clifft's 2^k active state is quoted).
Reproduce: clifft_env/bin/python reports/per_step_active_state/detailed_table_generate.py
"""
import csv
import glob
import math
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "reports/per_step_active_state"
CIRCS = ['coherent_d3_r1', 'coherent_d3_r3', 'coherent_d5_r1', 'coherent_d5_r5',
         'distillation', 'cultivation_d3', 'cultivation_d5', 'surface_d7_r7']
BYTES_PER_AMP = 16            # complex128


def load(circ):
    """Return per-circuit stats from the fused per-step trace."""
    f = f"{OUT}/fused_va_{circ}_per_step.csv"
    na, tr, res = [], [], []
    for r in csv.DictReader(open(f)):
        na.append(int(r['n_active']))
        tr.append(int(r['fused_transient_qubits']))
        res.append(int(r['fused_resident_qubits']))
    k = max(na)                                  # Clifft PEAK active rank == clifft_k
    return {
        'k': k, 'tr_pk': max(tr), 'res_pk': max(res),
        'clifft_sum': sum(1 << x for x in na),   # time-integrated dense active state
        'tr_sum': sum(1 << x for x in tr),
        'res_sum': sum(1 << x for x in res),
    }


def p2(x):                                       # integer exponent -> "2^x"
    return f"2^{x}"


def p2f(v):                                      # value -> "2^x.xx" (x = log2 v)
    if v <= 0:
        return "0"
    return f"2^{math.log2(v):.2f}"


def winx(num, den):                              # Clifft / fused  (>1 = fused wins)
    if den <= 0:
        return "inf"
    r = num / den
    if abs(r - 1.0) < 1e-9:
        return "parity"
    return f"{r:.3g}x" if r > 1 else f"{r:.3g}x (loss)"


def hbytes(b):
    for u, d in [('GiB', 1 << 30), ('MiB', 1 << 20), ('KiB', 1 << 10)]:
        if b >= d:
            return f"{b / d:.1f}{u}"
    return f"{b:.0f}B"


# ---------------------------------------------------------------- gather
S = {c: load(c) for c in CIRCS}

# ---------------------------------------------------------------- workbook
wb = openpyxl.Workbook()
HDR = Font(bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="305496")
GRPFILL = PatternFill("solid", fgColor="8EA9DB")
BOLD = Font(bold=True)
WINCOL = Font(color="1F7A1F")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center")


def style_header(ws, row, ncol):
    for j in range(1, ncol + 1):
        c = ws.cell(row=row, column=j)
        c.font = HDR; c.fill = HDRFILL; c.alignment = CTR; c.border = BORDER


def autofit(ws):
    for j in range(1, ws.max_column + 1):
        w = 0
        for i in range(1, ws.max_row + 1):
            v = ws.cell(row=i, column=j).value
            if v is not None:
                w = max(w, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = w + 3


# ===== Sheet 1: Transient vs Resident =====
ws = wb.active
ws.title = "Transient vs Resident"
hdr = ['circuit', 'Clifft PEAK 2^k', 'fused transient 2^b', 'transient vs Clifft',
       'fused resident 2^b', 'resident vs Clifft']
ws.append(hdr); style_header(ws, 1, len(hdr))
for c in CIRCS:
    s = S[c]
    ws.append([c, p2(s['k']), p2(s['tr_pk']), winx(1 << s['k'], 1 << s['tr_pk']),
               p2(s['res_pk']), winx(1 << s['k'], 1 << s['res_pk'])])
for r in range(2, 2 + len(CIRCS)):
    for j in range(1, len(hdr) + 1):
        ws.cell(row=r, column=j).border = BORDER
    ws.cell(row=r, column=4).font = WINCOL
    ws.cell(row=r, column=6).font = WINCOL
autofit(ws)


# ===== generic PEAK+SUM sheet builder (sheets 2 & 3) =====
def peak_sum_sheet(title, fmt, val):
    """fmt: scalar formatter; val(stat, key) returns the raw numeric for ratios.
    Columns: circuit | PEAK[Clifft, fused transient, x, fused resident, x] |
                       SUM [Clifft, fused transient, x, fused resident, x]."""
    ws = wb.create_sheet(title)
    # group header row
    ws.append([None, 'PEAK (max over steps)', None, None, None,
               'SUM (integrated over steps)', None, None, None])
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
    for j in (2, 6):
        cc = ws.cell(row=1, column=j); cc.font = BOLD; cc.fill = GRPFILL; cc.alignment = CTR
    sub = ['circuit',
           'Clifft', 'fused transient', 'x', 'fused resident', 'x',
           'Clifft', 'fused transient', 'x', 'fused resident', 'x']
    # fix: build header with 11 columns
    ws.delete_rows(1, 1)
    ws.append([None, 'PEAK (max over steps)', None, None, None, None,
               'SUM (integrated over steps)', None, None, None, None])
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
    for j in (2, 7):
        cc = ws.cell(row=1, column=j); cc.font = BOLD; cc.fill = GRPFILL; cc.alignment = CTR
    ws.append(sub); style_header(ws, 2, len(sub))
    for c in CIRCS:
        s = S[c]
        ck_pk = val(s, 'clifft_pk'); tr_pk = val(s, 'tr_pk'); rs_pk = val(s, 'res_pk')
        ck_sm = val(s, 'clifft_sum'); tr_sm = val(s, 'tr_sum'); rs_sm = val(s, 'res_sum')
        ws.append([c,
                   fmt(ck_pk), fmt(tr_pk), winx(ck_pk, tr_pk), fmt(rs_pk), winx(ck_pk, rs_pk),
                   fmt(ck_sm, sumlog=True), fmt(tr_sm, sumlog=True), winx(ck_sm, tr_sm),
                   fmt(rs_sm, sumlog=True), winx(ck_sm, rs_sm)])
    for r in range(3, 3 + len(CIRCS)):
        for j in range(1, len(sub) + 1):
            ws.cell(row=r, column=j).border = BORDER
        for j in (4, 6, 9, 11):
            ws.cell(row=r, column=j).font = WINCOL
    autofit(ws)
    return ws


# raw-value accessor in dense-equivalent dimension (amplitudes)
def dim_val(s, key):
    return {'clifft_pk': 1 << s['k'], 'tr_pk': 1 << s['tr_pk'], 'res_pk': 1 << s['res_pk'],
            'clifft_sum': s['clifft_sum'], 'tr_sum': s['tr_sum'], 'res_sum': s['res_sum']}[key]


def byte_val(s, key):
    return dim_val(s, key) * BYTES_PER_AMP


# Sheet 2: Active-State (2^x)
def fmt_2x(v, sumlog=False):
    return p2f(v) if sumlog else (p2(int(round(math.log2(v)))) if v > 0 else "0")
peak_sum_sheet("Active-State (2^x)", fmt_2x, dim_val)

# Sheet 3: Memory (bytes)
def fmt_bytes(v, sumlog=False):
    return hbytes(v)
peak_sum_sheet("Memory (bytes)", fmt_bytes, byte_val)

wb.save(f"{OUT}/DETAILED_TABLE.xlsx")
print("WROTE", f"{OUT}/DETAILED_TABLE.xlsx")


# ---------------------------------------------------------------- markdown companion
def md_table(rows):
    out = []
    out.append('| ' + ' | '.join(rows[0]) + ' |')
    out.append('|' + '|'.join('---' for _ in rows[0]) + '|')
    for r in rows[1:]:
        out.append('| ' + ' | '.join(r) + ' |')
    return '\n'.join(out)


lines = []
lines.append("# DETAILED TABLE — Clifft baseline vs live fused virtual-axis backend\n")
lines.append("All fused numbers are read from the per-step traces the dense-free single-frame "
             "`FusedSingleFrame` backend emits during its OWN run "
             "(`fused_va_<circuit>_per_step.csv`) — no TTN, no block, no Clifft state, no forced "
             "outcomes.  **state size** = active-state dimension `2^b` (the dense magic register); "
             "**memory** = `16·2^b` bytes (complex128); the `O(n^2)`-bit CHP tableau is excluded "
             "(poly, same basis as Clifft's `2^k`).  `transient` = peak fused workspace during a "
             "measurement-core contraction (`2^(W-1)`); `resident` = settled magic rank between "
             "measurements.\n")

lines.append("## 1. Transient vs Resident (PEAK active-state)\n")
r = [['circuit', 'Clifft PEAK 2^k', 'fused transient 2^b', 'transient vs Clifft',
      'fused resident 2^b', 'resident vs Clifft']]
for c in CIRCS:
    s = S[c]
    r.append([c, p2(s['k']), p2(s['tr_pk']), winx(1 << s['k'], 1 << s['tr_pk']),
              p2(s['res_pk']), winx(1 << s['k'], 1 << s['res_pk'])])
lines.append(md_table(r) + "\n")

lines.append("## 2. Active-State (2^x) — PEAK and integrated SUM\n")
r = [['circuit', 'Clifft PEAK', 'fused tr. PEAK', 'tr x', 'fused res. PEAK', 'res x',
      'Clifft SUM', 'fused tr. SUM', 'tr x', 'fused res. SUM', 'res x']]
for c in CIRCS:
    s = S[c]
    r.append([c, fmt_2x(dim_val(s, 'clifft_pk')), fmt_2x(dim_val(s, 'tr_pk')),
              winx(dim_val(s, 'clifft_pk'), dim_val(s, 'tr_pk')),
              fmt_2x(dim_val(s, 'res_pk')), winx(dim_val(s, 'clifft_pk'), dim_val(s, 'res_pk')),
              fmt_2x(s['clifft_sum'], sumlog=True), fmt_2x(s['tr_sum'], sumlog=True),
              winx(s['clifft_sum'], s['tr_sum']),
              fmt_2x(s['res_sum'], sumlog=True), winx(s['clifft_sum'], s['res_sum'])])
lines.append(md_table(r) + "\n")

lines.append("## 3. Memory (bytes) — PEAK and integrated SUM\n")
r = [['circuit', 'Clifft PEAK', 'fused tr. PEAK', 'tr x', 'fused res. PEAK', 'res x',
      'Clifft SUM', 'fused tr. SUM', 'tr x', 'fused res. SUM', 'res x']]
for c in CIRCS:
    s = S[c]
    r.append([c, hbytes(byte_val(s, 'clifft_pk')), hbytes(byte_val(s, 'tr_pk')),
              winx(byte_val(s, 'clifft_pk'), byte_val(s, 'tr_pk')),
              hbytes(byte_val(s, 'res_pk')), winx(byte_val(s, 'clifft_pk'), byte_val(s, 'res_pk')),
              hbytes(byte_val(s, 'clifft_sum')), hbytes(byte_val(s, 'tr_sum')),
              winx(byte_val(s, 'clifft_sum'), byte_val(s, 'tr_sum')),
              hbytes(byte_val(s, 'res_sum')), winx(byte_val(s, 'clifft_sum'), byte_val(s, 'res_sum'))])
lines.append(md_table(r) + "\n")

open(f"{OUT}/DETAILED_TABLE.md", "w").write('\n'.join(lines))
print("WROTE", f"{OUT}/DETAILED_TABLE.md")
