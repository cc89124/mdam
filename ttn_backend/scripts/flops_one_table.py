"""Render ONE combined FLOP table: per-operation FLOPs + TOTAL + advantage vs Clifft.

Values are the measured/modelled SUM-over-run results from
per_step_flops_compare.py (FLOP unit; complex mult=6, add=2, norm=4, vdot=8).
Each backend's TOTAL = its FLOP ops + the shared Clifford bit-op floor, so the
TOTAL column is the complete operation count. Advantage = Clifft TOTAL / TOTAL.
"""
import os
from pathlib import Path

OUT = Path("reports/per_step_flops")

# circuit: (bitop, clifft_mm, ttn_contract, ttn_qr, nc_mm, nc_norm, ttn_trunc)
# ttn_contract/qr = None when TTN failed to lay out (surface_d7_r7).
DATA = {
    "coherent_d3_r1":  (3.5e3,  17.3e3,  12.1e3, 11.3e3, 0.0,     0.0,     False),
    "coherent_d3_r3":  (19.1e3, 545.2e3, 1.2e6,  1.0e6,  68.1e3,  132.0e3, False),
    "coherent_d5_r1":  (70.6e3, 12.7e6,  7.9e6,  7.6e6,  0.0,     0.0,     False),
    "coherent_d5_r5":  (1.2e6,  209.3e9, 8.2e12, 8.9e12, 799.3e6, 2.024e9, True),
    "distillation":    (12.5e3, 12.6e3,  8.8e3,  6.6e3,  2.852e3, 2.296e3, False),
    "cultivation_d3":  (2.8e3,  23.6e3,  75.9e3, 58.1e3, 11.49e3, 20.86e3, False),
    "cultivation_d5":  (73.7e3, 3.4e6,   28.2e6, 19.5e6, 1.94e6,  10.03e6, False),
    "surface_d7_r7":   (0.0,    4.7e3,   None,   None,   0.0,     0.0,     False),
}

HEAD = ["circuit", "Clifford bit-op",
        "Clifft matmul", "Clifft TOTAL",
        "TTN contract", "TTN QR", "TTN TOTAL", "TTN x",
        "NC matmul", "NC norm", "NC TOTAL", "NC x"]


def hf(n):
    if n is None:
        return "-"
    n = float(n)
    for u in ("", "K", "M", "G", "T", "P"):
        if abs(n) < 1000 or u == "P":
            return f"{n:.1f}{u}" if u else f"{n:.0f}"
        n /= 1000


def adv(base, x):
    if x is None or base is None:
        return "-"
    if x == 0:
        return "∞" if base > 0 else "-"
    r = base / x
    return f"{r:.0f}x" if r >= 100 else (f"{r:.1f}x" if r >= 1 else f"{r:.2f}x")


def rows():
    out = []
    for c, (bit, cl, ct, qr, ncm, ncn, tr) in DATA.items():
        clt = cl + bit
        ttt = (ct + qr + bit) if ct is not None else None
        nct = ncm + ncn + bit
        fl = "†" if tr else ""
        out.append([
            c, hf(bit),
            hf(cl), hf(clt),
            (hf(ct) + fl) if ct is not None else "-", hf(qr),
            (hf(ttt) + fl) if ttt is not None else "-", adv(clt, ttt),
            hf(ncm), hf(ncn), hf(nct), adv(clt, nct),
        ])
    return out


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    R = rows()
    # markdown
    md = ["# Total compute by operation: Clifft (baseline) vs TTN vs near-Clifford\n",
          "FLOP unit (complex mult=6, add=2, norm=4, vdot=8), SUM over the whole run. "
          "Per-operation columns + **TOTAL** (= the backend's FLOP ops + the shared "
          "Clifford bit-op floor) + advantage `x` = Clifft TOTAL / backend TOTAL. "
          "Clifft = analytic 2^k dense model (matmul only); TTN & near-Clifford = "
          "MEASURED. SVD = 0 (TTN exact mode). `†` coherent_d5_r5 TTN = executed prefix "
          "(~step 2289). surface_d7_r7 TTN fails to lay out -> '-'.\n",
          "Clifford bit-op = polynomial GF(2) tableau/frame work (gate~n, meas~n^2, "
          "deferred rot~n); it is why near-Clifford's `0 FLOP` cases are bit-ops-only, "
          "not no-work.\n\n",
          "> **Compute is a different axis from memory — and frame reduction helps "
          "both.** With frame reduction ON (default), peeling each measured-out qubit's "
          "dead residue keeps the magic blocks smaller, so the `norm`/`vdot` **factoring "
          "scan** runs over far fewer amplitudes: NC FLOP drops sharply vs the "
          "pre-reduction numbers (`cultivation_d5` `150M→12M`, `distillation` `41K→18K`, "
          "`coherent_d5_r5` norm `12.5G→2.0G`). This flips `distillation` to a **compute "
          "win (1.4x)** and lifts every coherent circuit (`d5_r5` 16x→**74x**). The "
          "remaining `NC x < 1x` rows (`cultivation_d3` 0.75x, `cultivation_d5` 0.29x, "
          "both up from 0.20x/0.02x) are *compute* losses, not memory losses: on **all-"
          "magic** circuits near-Clifford still pays a factoring scan on genuinely-"
          "irreducible magic that Clifft's analytic `2^k` model never spends, so it does "
          "more FLOP even when its memory is parity/better — the expected trade for the "
          "bounded block (§8.3/§8.4).\n\n",
          "| " + " | ".join(HEAD) + " |",
          "|" + "---|" + "--:|" * (len(HEAD) - 1)]
    for r in R:
        md.append("| " + " | ".join(r) + " |")
    text = "\n".join(md) + "\n"
    (OUT / "FLOPS_TABLE.md").write_text(text)
    print(text)
    # excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "FLOP by operation"
    head = PatternFill("solid", fgColor="305496"); hfont = Font(bold=True, color="FFFFFF")
    tot = PatternFill("solid", fgColor="FFF2CC"); advf = PatternFill("solid", fgColor="E2EFDA")
    extra = PatternFill("solid", fgColor="FCE4D6"); bit = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="BFBFBF"); bd = Border(thin, thin, thin, thin)
    ws.append(HEAD)
    for r in R:
        ws.append(r)
    for col in range(1, len(HEAD) + 1):
        cc = ws.cell(row=1, column=col)
        cc.fill = head; cc.font = hfont
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = bd
    TOT_COLS = {4, 7, 11}; ADV_COLS = {8, 12}; EXTRA_COLS = {6, 10}; BIT_COLS = {2}
    for ri in range(2, 2 + len(R)):
        for col in range(1, len(HEAD) + 1):
            cell = ws.cell(row=ri, column=col)
            cell.border = bd
            cell.alignment = Alignment(horizontal="right") if col > 1 else Alignment(horizontal="left")
            if col in ADV_COLS:
                cell.fill = advf; cell.font = Font(bold=True)
            elif col in TOT_COLS:
                cell.fill = tot; cell.font = Font(bold=True)
            elif col in EXTRA_COLS:
                cell.fill = extra
            elif col in BIT_COLS:
                cell.fill = bit
    ws.column_dimensions["A"].width = 18
    for col in range(2, len(HEAD) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "B2"
    wb.save(OUT / "FLOPS_TABLE.xlsx")
    print(f"Excel written: {OUT / 'FLOPS_TABLE.xlsx'}")


if __name__ == "__main__":
    main()
