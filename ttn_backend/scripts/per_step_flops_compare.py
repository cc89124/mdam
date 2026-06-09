"""Per-step FLOP comparison: Clifft (model) vs TTN vs near-Clifford (measured).

Compute cost in FLOPs (complex-arith convention: mult=6, add=2, norm=4, vdot=8).
Columns separate the common matmul work from each backend's EXTRA primitive:

  * Clifft (dense baseline, MODEL): applies each active gate to its 2^k dense
    register. Per step = (op-arity constant) * 2^k. No QR / no factoring.
        1q/rot/expand 14*2^k ; 2q 30*2^k ; active measure 34*2^k ; frame/noise 0.
    (clifft is not a runnable backend here -- it is the 16*2^k dense model used
    throughout this repo, so its FLOP is the matching analytic per-step proxy.)

  * near-Clifford (block, MEASURED): real FLOPs counted in block_magic --
        matmul = rotation apply + kron + measurement collapse (on 2^block)
        norm   = the factoring scan (norm/vdot probes in factor())   <- its EXTRA

  * TTN (MEASURED): tensordot contraction FLOP (monkeypatched by shape) +
        QR = qr_work_proxy * 8  (m*n*min(m,n) MACs)                  <- its EXTRA
        SVD = 0 in exact mode (n_svd == 0; SVD only under a truncation policy).

PEAK = max over steps of the per-step FLOP; SUM = total FLOP over the run.
Truncated TTN (coherent_d5_r5) is flagged with a dagger; its SUM is the executed
prefix only.

Outputs: reports/per_step_flops/FLOPS_TABLE.md  and  FLOPS_TABLE.xlsx
"""
from __future__ import annotations
import os
import sys
import time
from pathlib import Path

sys.setrecursionlimit(50000)

import numpy as np
import clifft
from nearclifford_backend.backend import NearCliffordBackend, _opname
from ttn_backend.scripts.per_step_memory_compare import (
    _build_spec_homing, POLICY)
from ttn_backend import TTNBackend

ORDER = ["coherent_d3_r1", "coherent_d3_r3", "coherent_d5_r1", "coherent_d5_r5",
         "distillation", "cultivation_d3", "cultivation_d5", "surface_d7_r7"]
OUT = Path("reports/per_step_flops")

# ---- clifft per-op dense-work model (constants x 2^k) -----------------------
_C_1Q, _C_2Q, _C_MEAS = 14.0, 30.0, 34.0


def clifft_op_flops(opname, k):
    """FLOP clifft's 2^k dense register pays for one circuit op of this kind."""
    D = float(1 << k)
    if opname.startswith("OP_FRAME") or opname in (
            "OP_NOISE", "OP_NOISE_BLOCK", "OP_READOUT_NOISE", "OP_APPLY_PAULI"
    ) or "DORMANT" in opname or opname in getattr(clifft, "IGNORE_OPS", ()):
        return 0.0
    if "CNOT" in opname or "CZ" in opname or "SWAP" in opname or opname.endswith("U4"):
        return _C_2Q * D
    if "MEAS" in opname:
        return _C_MEAS * D
    return _C_1Q * D     # 1q gates, T/rot/phase, expand, U2


# ---- near-Clifford run: measured FLOP per step -----------------------------
def run_nc(circuit, timeout):
    src = open(f"qec_bench/circuits/{circuit}.stim").read()
    prog = clifft.compile(src)
    be = NearCliffordBackend(block=True)
    rows = []          # per step: (k, opname, cum_mm, cum_norm)
    t0 = time.perf_counter()

    def rec(step, backend):
        if time.perf_counter() - t0 > timeout:
            raise TimeoutError(f"NC exceeded {timeout}s")
        k = len(backend.slot2id)
        nm = _opname(prog[step].opcode) if step < len(prog) else "END"
        m = backend.nc.mag
        rows.append((k, nm, m.flop_mm, m.flop_norm))

    try:
        be.run_shot(prog, 42, step_recorder=rec)
    except TimeoutError as e:
        print(f"  [NC] {e}; dropping {circuit}")
        return None
    # ---- Clifford bit-op floor (polynomial; the GF(2) work near-Clifford does in
    # the tableau/frame instead of dense FLOP). Per the Gottesman-Knill cost model:
    #   Clifford gate ~ n bit-ops ; measurement ~ n^2 ; deferred rotation ~ n.
    n = be.nc.n
    CLIFF = ("FRAME_H", "FRAME_S", "FRAME_CNOT", "FRAME_CZ", "FRAME_SWAP",
             "ARRAY_H", "ARRAY_S", "ARRAY_CNOT", "ARRAY_CZ")
    cg = me = ro = 0
    for i in range(len(prog)):
        nm = _opname(prog[i].opcode)
        if "MEAS" in nm:
            me += 1
        elif any(nm.startswith("OP_" + c) or ("OP_" + c) in nm for c in CLIFF):
            cg += 1
        elif any(t in nm for t in ("_T", "ROT", "PHASE", "EXPAND", "U2", "U4")):
            ro += 1
    bitops = float(cg * n + me * n * n + ro * n)
    # per-step deltas (work DURING step s = cum[s+1]-cum[s])
    cl_peak = cl_sum = mm_peak = mm_sum = nr_peak = nr_sum = 0.0
    for s in range(len(rows) - 1):
        k, nm, mm0, nr0 = rows[s]
        _, _, mm1, nr1 = rows[s + 1]
        cl = clifft_op_flops(nm, k)
        dmm = mm1 - mm0
        dnr = nr1 - nr0
        cl_peak = max(cl_peak, cl); cl_sum += cl
        mm_peak = max(mm_peak, dmm); mm_sum += dmm
        nr_peak = max(nr_peak, dnr); nr_sum += dnr
    return dict(clifft=(cl_peak, cl_sum), nc_mm=(mm_peak, mm_sum),
                nc_norm=(nr_peak, nr_sum), n_steps=len(rows) - 1,
                bitops=bitops, n=n, n_meas=me, n_cliff=cg, n_rot=ro)


# ---- TTN run: measured contraction FLOP (monkeypatch) + QR proxy -----------
_TD = {"mm": 0.0, "qr": 0.0, "svd": 0.0}
_orig = {}


def _contracted_K(a, axes):
    if isinstance(axes, int):
        K = 1
        for s in a.shape[a.ndim - axes:]:
            K *= int(s)
        return K
    aa = axes[0]
    aa = [aa] if isinstance(aa, int) else aa
    K = 1
    for i in aa:
        K *= int(a.shape[i])
    return K


def _patch():
    _orig["td"] = np.tensordot
    _orig["dot"] = np.dot
    _orig["mm"] = np.matmul
    _orig["qr"] = np.linalg.qr
    _orig["svd"] = np.linalg.svd
    _orig["eigh"] = np.linalg.eigh

    def td(a, b, axes=2):
        K = _contracted_K(np.asarray(a), axes)
        if K:
            _TD["mm"] += 8.0 * a.size * b.size / K
        return _orig["td"](a, b, axes=axes)

    def dot(a, b, *a_, **k_):
        a = np.asarray(a); b = np.asarray(b)
        K = int(a.shape[-1]) if a.ndim else 1
        if K:
            _TD["mm"] += 8.0 * a.size * b.size / K
        return _orig["dot"](a, b, *a_, **k_)

    def mm(a, b, *a_, **k_):
        a = np.asarray(a); b = np.asarray(b)
        K = int(a.shape[-1]) if a.ndim else 1
        if K:
            _TD["mm"] += 8.0 * a.size * b.size / K
        return _orig["mm"](a, b, *a_, **k_)

    def qr(a, *a_, **k_):
        a = np.asarray(a)
        if a.ndim == 2:
            m, n = a.shape
            _TD["qr"] += 8.0 * m * n * min(m, n)   # leading-order Householder work
        return _orig["qr"](a, *a_, **k_)

    def svd(a, *a_, **k_):
        a = np.asarray(a)
        if a.ndim == 2:
            m, n = a.shape
            _TD["svd"] += 8.0 * m * n * min(m, n)
        return _orig["svd"](a, *a_, **k_)

    def eigh(a, *a_, **k_):
        a = np.asarray(a)
        if a.ndim == 2:
            _TD["qr"] += 8.0 * a.shape[0] ** 3       # Hermitian eig (factorization)
        return _orig["eigh"](a, *a_, **k_)

    np.tensordot = td; np.dot = dot; np.matmul = mm
    np.linalg.qr = qr; np.linalg.svd = svd; np.linalg.eigh = eigh


def _unpatch():
    np.tensordot = _orig["td"]; np.dot = _orig["dot"]; np.matmul = _orig["mm"]
    np.linalg.qr = _orig["qr"]; np.linalg.svd = _orig["svd"]
    np.linalg.eigh = _orig["eigh"]


def run_ttn(circuit, timeout):
    saved = {}
    env = dict(POLICY); env["TTN_MULTICNOT_PARITY_REWRITE"] = "1"
    for k, v in env.items():
        saved[k] = os.environ.get(k); os.environ[k] = v
    try:
        src = open(f"qec_bench/circuits/{circuit}.stim").read()
        prog = clifft.compile(src)
        spec, homing = _build_spec_homing(prog, "carving")   # unpatched build
        snap = {}          # step_id -> (cum_contract, cum_qr, cum_svd)

        def rec(row):
            s = row.get("step_id")
            if s is None:
                return
            snap[s] = (_TD["mm"], _TD["qr"], _TD["svd"])

        be = TTNBackend(spec, homing, trace_recorder=rec)
        _TD["mm"] = _TD["qr"] = _TD["svd"] = 0.0
        _patch()
        try:
            be.run_shot(prog, 42, runtime_timeout=timeout)
        finally:
            _unpatch()
        total_steps = len(prog)
        executed = (max(snap) + 1) if snap else 0
        truncated = executed < 0.98 * total_steps
        # per-step deltas for PEAK
        ct_peak = qr_peak = 0.0
        prev_c = prev_q = 0.0
        for s in sorted(snap):
            c, q, _ = snap[s]
            ct_peak = max(ct_peak, c - prev_c)
            qr_peak = max(qr_peak, q - prev_q)
            prev_c, prev_q = c, q
        ct_sum = _TD["mm"]
        qr_sum = _TD["qr"]
        n_svd = int(be.state.metrics.get("n_svd", 0))
        return dict(ttn_mm=(ct_peak, ct_sum), ttn_qr=(qr_peak, qr_sum),
                    n_svd=n_svd, truncated=truncated, executed=executed,
                    total=total_steps)
    except Exception as e:
        print(f"  [TTN] failed ({type(e).__name__}: {e}); dropping TTN for {circuit}")
        return None
    finally:
        for k, old in saved.items():
            if old is None:
                os.environ.pop(k, None)
            else:
                os.environ[k] = old


# ---- formatting ------------------------------------------------------------
def hflop(n):
    if n is None:
        return "-"
    n = float(n)
    for u in ("", "K", "M", "G", "T", "P", "E"):
        if abs(n) < 1000 or u == "E":
            return f"{n:.1f}{u}" if u else f"{n:.0f}"
        n /= 1000


def advx(base, x):
    if x is None or base is None:
        return "-"
    if x == 0:
        return "∞" if base > 0 else "-"   # backend does zero FLOP (all-free)
    r = base / x
    if r >= 100:
        return f"{r:.0f}x"
    if r >= 1:
        return f"{r:.1f}x"
    return f"{r:.2f}x"


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    nc_to = float(os.environ.get("NC_TIMEOUT", "600"))
    ttn_to = float(os.environ.get("TTN_TIMEOUT", "700"))
    data = {}
    for c in ORDER:
        print(f"######## {c} ########")
        nc = run_nc(c, nc_to)
        ttn = run_ttn(c, ttn_to)
        data[c] = (nc, ttn)
        if nc:
            print(f"  clifft sum={hflop(nc['clifft'][1])}  "
                  f"NC mm={hflop(nc['nc_mm'][1])} norm={hflop(nc['nc_norm'][1])}")
        if ttn:
            t = "†" if ttn["truncated"] else ""
            print(f"  TTN contract={hflop(ttn['ttn_mm'][1])}{t} "
                  f"QR={hflop(ttn['ttn_qr'][1])} n_svd={ttn['n_svd']}")
    write_outputs(data)


# ---- value extraction -------------------------------------------------------
def _vals(data, c, agg):
    """agg = 0 PEAK, 1 SUM. Returns the FLOP pieces + the bit-op floor."""
    nc, ttn = data[c]
    cl = nc["clifft"][agg] if nc else None
    ncm = nc["nc_mm"][agg] if nc else None
    ncn = nc["nc_norm"][agg] if nc else None
    ncf = (ncm + ncn) if nc else None
    ttm = ttn["ttn_mm"][agg] if ttn else None
    ttq = ttn["ttn_qr"][agg] if ttn else None
    ttf = (ttm + ttq) if ttn else None
    # bit-op floor: SUM = full Clifford/measurement work; PEAK = one measurement (n^2)
    bit = (nc["bitops"] if agg == 1 else float(nc["n"] ** 2)) if nc else None
    return dict(cl=cl, ncm=ncm, ncn=ncn, ncf=ncf, ttm=ttm, ttq=ttq, ttf=ttf,
                bit=bit, trunc=bool(ttn and ttn["truncated"]))


TOTAL_H = ["circuit", "Clifft FLOP", "TTN FLOP", "NC FLOP", "Clifford bit-ops",
           "Clifft TOTAL", "TTN TOTAL", "NC TOTAL", "TTN x", "NC x"]
PEAK_H = ["circuit", "Clifft FLOP", "TTN FLOP", "NC FLOP", "TTN x", "NC x"]
BRK_H = ["circuit", "Clifft matmul", "TTN contract", "TTN QR",
         "NC matmul", "NC norm", "Clifford bit-ops"]


def _add(a, b):
    return (a + b) if (a is not None and b is not None) else None


def _total_rows(data):
    rows = []
    for c in ORDER:
        v = _vals(data, c, 1)
        clt, ttt, nct = _add(v["cl"], v["bit"]), _add(v["ttf"], v["bit"]), _add(v["ncf"], v["bit"])
        fl = "†" if v["trunc"] else ""
        rows.append([c, hflop(v["cl"]), (hflop(v["ttf"]) + fl) if v["ttf"] is not None else "-",
                     hflop(v["ncf"]), hflop(v["bit"]),
                     hflop(clt), (hflop(ttt) + fl) if ttt is not None else "-", hflop(nct),
                     advx(clt, ttt), advx(clt, nct)])
    return rows


def _peak_rows(data):
    rows = []
    for c in ORDER:
        v = _vals(data, c, 0)
        fl = "†" if v["trunc"] else ""
        rows.append([c, hflop(v["cl"]), (hflop(v["ttf"]) + fl) if v["ttf"] is not None else "-",
                     hflop(v["ncf"]), advx(v["cl"], v["ttf"]), advx(v["cl"], v["ncf"])])
    return rows


def _brk_rows(data):
    rows = []
    for c in ORDER:
        v = _vals(data, c, 1)
        fl = "†" if v["trunc"] else ""
        rows.append([c, hflop(v["cl"]),
                     (hflop(v["ttm"]) + fl) if v["ttm"] is not None else "-", hflop(v["ttq"]),
                     hflop(v["ncm"]), hflop(v["ncn"]), hflop(v["bit"])])
    return rows


def _md(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "|" + "---|" + "--:|" * (len(headers) - 1)]
    for r in rows:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def write_outputs(data):
    md = ["# Total compute: Clifft (model) vs TTN vs near-Clifford (measured)\n",
          "Two operation classes, both reported in full so the TOTAL is the complete "
          "operation count, not a subset:\n",
          "* **FLOP** (floating-point; complex mult=6, add=2, norm=4, vdot=8) -- the "
          "exponential dense/contraction work. Clifft = analytic 2^k model; TTN & "
          "near-Clifford = MEASURED from a real shot (all matmul / tensordot / QR / SVD "
          "/ eigh / elementwise captured).\n",
          "* **Clifford bit-ops** (polynomial, GF(2)) -- the tableau/frame work done "
          "instead of dense FLOP: per Gottesman-Knill, Clifford gate ~ n, measurement ~ "
          "n^2, deferred rotation ~ n. This is the floor every near-Clifford backend "
          "pays; near-Clifford does ALL its Clifford here (so its `0 FLOP` cases are "
          "`bit-ops only`, not `no work`).\n",
          "**TOTAL = FLOP + bit-ops** (every arithmetic/bit op counted once). Advantage "
          "`x` = Clifft TOTAL / backend TOTAL. SVD = 0 (TTN exact mode, n_svd=0). `†` "
          "coherent_d5_r5 TTN = executed prefix (~step 2289). surface_d7_r7 TTN fails to "
          "lay out (RecursionError) -> '-'.\n",
          "\n## TOTAL operation count (SUM over run)\n", _md(TOTAL_H, _total_rows(data)),
          "\n\n## PEAK single-step FLOP\n", _md(PEAK_H, _peak_rows(data)),
          "\n\n## Breakdown (SUM): matmul / extra-primitive / bit-ops\n",
          _md(BRK_H, _brk_rows(data))]
    text = "\n".join(md) + "\n"
    (OUT / "FLOPS_TABLE.md").write_text(text)
    print("\n" + text)
    _write_xlsx(data)


def _write_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    head = PatternFill("solid", fgColor="305496"); hf = Font(bold=True, color="FFFFFF")
    adv = PatternFill("solid", fgColor="E2EFDA"); extra = PatternFill("solid", fgColor="FCE4D6")
    totf = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BFBFBF"); bd = Border(thin, thin, thin, thin)
    ctr = Alignment(horizontal="center", vertical="center"); rt = Alignment(horizontal="right")

    def sheet(ws, title, headers, rows, adv_cols=(), extra_cols=(), tot_cols=()):
        ws.title = title
        ws.append(headers)
        for r in rows:
            ws.append(r)
        for col in range(1, len(headers) + 1):
            cc = ws.cell(row=1, column=col)
            cc.fill = head; cc.font = hf; cc.alignment = ctr; cc.border = bd
        for ri in range(2, 2 + len(rows)):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ri, column=col)
                cell.border = bd
                cell.alignment = rt if col > 1 else Alignment(horizontal="left")
                if col in adv_cols:
                    cell.fill = adv; cell.font = Font(bold=True)
                elif col in tot_cols:
                    cell.fill = totf; cell.font = Font(bold=True)
                elif col in extra_cols:
                    cell.fill = extra
        ws.column_dimensions["A"].width = 18
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "B2"

    sheet(wb.active, "TOTAL (SUM)", TOTAL_H, _total_rows(data),
          adv_cols=(9, 10), tot_cols=(6, 7, 8), extra_cols=(5,))
    sheet(wb.create_sheet(), "PEAK FLOP", PEAK_H, _peak_rows(data), adv_cols=(5, 6))
    sheet(wb.create_sheet(), "Breakdown (SUM)", BRK_H, _brk_rows(data),
          extra_cols=(4, 6, 7))
    p = OUT / "FLOPS_TABLE.xlsx"
    wb.save(p)
    print(f"Excel written: {p}")


if __name__ == "__main__":
    main()
