"""Detailed active-state-size + memory table from the per-step CSVs.

Two metrics, each with PEAK and SUM combined into ONE table, baseline = Clifft:
  * ACTIVE-STATE SIZE = dense-equivalent dimension, written in 2^ form:
        clifft 2^k ; TTN 2^log2(stored/16) ; near-Clifford 2^max_magic_block
  * MEMORY (bytes): clifft 16*2^k ; TTN stored bytes ; near-Clifford magic+tableau+pending
  * PEAK = max over steps ; SUM = area under the per-step curve (sum over steps).
Advantage vs clifft = clifft / backend, reported as a MULTIPLE only (no percent).

Truncated TTN (coherent_d5_r5 stops at step ~2289): its PEAK is still its true
peak; its SUM is over the executed prefix only (flagged with a dagger), and the
TTN SUM ratio uses the Clifft sum over that SAME prefix.

Outputs:  reports/per_step_active_state/DETAILED_TABLE.md  (markdown)
          reports/per_step_active_state/DETAILED_TABLE.xlsx (Excel)
"""
from __future__ import annotations
import csv
import math
import os

ORDER = ["coherent_d3_r1", "coherent_d3_r3", "coherent_d5_r1", "coherent_d5_r5",
         "distillation", "cultivation_d3", "cultivation_d5", "surface_d7_r7"]
SRC = "reports/per_step_active_state"

HEADERS = ["circuit",
           "Clifft PEAK", "TTN PEAK", "TTN x", "near-Clifford PEAK", "NC x",
           "Clifft SUM", "TTN SUM", "TTN x", "near-Clifford SUM", "NC x"]


def load(circ):
    rows = list(csv.DictReader(open(os.path.join(SRC, f"{circ}_per_step.csv"))))
    cl_mem, nc_mem, k, ncq = [], [], [], []
    tt_mem = []                       # None where TTN absent/truncated
    for r in rows:
        k.append(int(r["n_active"]))
        ncq.append(int(r["near_clifft_qubits"]) if r["near_clifft_qubits"] != "" else 0)
        cl_mem.append(int(r["clifft_dense_bytes"]))
        nc_mem.append(int(r["near_clifft_bytes"]) if r["near_clifft_bytes"] != "" else 0)
        tt_mem.append(int(r["ttn_stored_bytes"]) if r["ttn_stored_bytes"] != "" else None)
    return cl_mem, tt_mem, nc_mem, k, ncq


def hbytes(n):
    if n is None:
        return "-"
    n = float(n)
    for u in ("B", "KiB", "MiB", "GiB", "TiB", "PiB"):
        if n < 1024 or u == "PiB":
            return f"{n:.1f}{u}"
        n /= 1024


def pow2(n):
    """Active-state size in 2^x form (x = log2 of the dimension)."""
    if n is None:
        return "-"
    if n <= 0:
        return "0"
    x = math.log2(n)
    xr = round(x, 2)
    if abs(xr - round(xr)) < 1e-9:
        return f"2^{int(round(xr))}"
    return f"2^{xr:.2f}"


def advx(base, x):
    """Advantage as a bare multiple: clifft / backend."""
    if x is None or x == 0 or base is None:
        return "-"
    r = base / x
    if r >= 100:
        return f"{r:.0f}x"
    if r >= 1:
        return f"{r:.1f}x"
    return f"{r:.2f}x"


def metric_rows(fmt, getcl, gettt, getnc):
    rows = []
    for c in ORDER:
        cl, tt, nc, k, ncq = load(c)
        clseq = getcl(cl, k, ncq)
        ttseq = gettt(tt, k, ncq)
        ncseq = getnc(nc, k, ncq)
        # PEAK
        cl_pk = max(clseq)
        present = [i for i in range(len(tt)) if tt[i] is not None]
        tt_pk = max(ttseq[i] for i in present) if present else None
        nc_pk = max(ncseq)
        # SUM
        cl_sum = sum(clseq)
        nc_sum = sum(ncseq)
        tt_sum = sum(ttseq[i] for i in present) if present else None
        cl_sum_tt = sum(clseq[i] for i in present) if present else None
        flag = "†" if (present and len(present) < len(tt)) else ""
        rows.append([
            c,
            fmt(cl_pk), fmt(tt_pk), advx(cl_pk, tt_pk), fmt(nc_pk), advx(cl_pk, nc_pk),
            fmt(cl_sum), (fmt(tt_sum) + flag) if tt_sum is not None else "-",
            advx(cl_sum_tt, tt_sum) if tt_sum is not None else "-",
            fmt(nc_sum), advx(cl_sum, nc_sum),
        ])
    return rows


def md_table(rows):
    out = ["| " + " | ".join(HEADERS) + " |"]
    align = ["---"] + ["--:"] * (len(HEADERS) - 1)
    out.append("| " + " | ".join(align) + " |")
    for r in rows:
        out.append("| " + " | ".join(str(c) for c in r) + " |")
    return "\n".join(out)


# --- RESIDENT (settled) no-regression: block factoring guarantees the settled
#     max_block never exceeds Clifft's active dimension. The PEAK columns above are
#     the intra-step TRANSIENT high-water (a measurement's core flush), which can
#     spike +1 from frame over-promotion (then the projector/purge removes it). This
#     table compares the SETTLED resident peak against Clifft -- the honest
#     no-regression statement. ---
NOREG_HEADERS = ["circuit", "Clifft PEAK 2^k",
                 "NC transient 2^b", "transient vs Clifft",
                 "NC resident 2^b", "resident vs Clifft"]


def load_peaks(circ):
    rows = list(csv.DictReader(open(os.path.join(SRC, f"{circ}_per_step.csv"))))
    k, ncqt, ncqr = [], [], []
    for r in rows:
        k.append(int(r["n_active"]))
        t = r.get("near_clifft_qubits", "")
        ncqt.append(int(t) if t != "" else 0)
        v = r.get("near_clifft_qubits_resident", "")
        ncqr.append(int(v) if v != "" else 0)
    return k, ncqt, ncqr


def _verdict(clpk, pk):
    if pk > clpk:
        return "LOSS (+%d)" % (pk - clpk)
    if pk == clpk:
        return "parity"
    return "%dx win" % (1 << (clpk - pk))


def noreg_rows():
    rows = []
    for c in ORDER:
        k, ncqt, ncqr = load_peaks(c)
        clpk = max(k) if k else 0
        tpk = max(ncqt) if ncqt else 0
        rpk = max(ncqr) if ncqr else 0
        rows.append([c, "2^%d" % clpk,
                     "2^%d" % tpk, _verdict(clpk, tpk),
                     "2^%d" % rpk, _verdict(clpk, rpk)])
    return rows


def md_noreg(rows):
    out = ["| " + " | ".join(NOREG_HEADERS) + " |"]
    out.append("| " + " | ".join(["---"] + ["--:"] * (len(NOREG_HEADERS) - 1)) + " |")
    for r in rows:
        out.append("| " + " | ".join(str(c) for c in r) + " |")
    return "\n".join(out)


ACTIVE = dict(getcl=lambda cl, k, ncq: [c // 16 for c in cl],                 # 2^k
              gettt=lambda tt, k, ncq: [(t // 16 if t is not None else None) for t in tt],
              getnc=lambda nc, k, ncq: [1 << q for q in ncq],                 # 2^max_block
              fmt=pow2)
MEMORY = dict(getcl=lambda cl, k, ncq: cl,
              gettt=lambda tt, k, ncq: tt,
              getnc=lambda nc, k, ncq: nc,
              fmt=hbytes)


def write_md(active, memory):
    out = []
    out.append("# Detailed per-step ACTIVE-STATE & MEMORY table\n")
    out.append("Baseline = **Clifft**. The `x` columns are the advantage = Clifft / backend "
               "(a bare multiple; >1x = backend is that many times smaller). PEAK and SUM "
               "are combined in one table per metric. Active-state size is written in `2^x` "
               "form (x = log2 of the dense-equivalent dimension). coherent_d7_r1/_d7_r7 "
               "excluded.\n")
    out.append("near-Clifford here is the **intra-step transient high-water mark** (the "
               "honest memory-feasibility peak: a measurement's anticommutation-core flush "
               "briefly forms a larger entangled block before its projector collapses it). "
               "The settled step-boundary **resident** value (lower; e.g. coherent_d5_r5 "
               "`2^12` resident vs `2^13` transient) is in `SUMMARY_TABLE.md`.\n")
    out.append("`†` coherent_d5_r5 is the full 3228-step circuit; its TTN line stops at step "
               "~2289 (full-circuit TTN does not finish), so the TTN **SUM** is over that "
               "prefix and the TTN SUM ratio uses the Clifft sum over the same prefix. "
               "surface_d7_r7 is frame-only (no active idents); TTN fails to lay out -> '-'.\n")
    out.append("> **Read the `<1x` cells honestly — `cultivation_d5` is the all-magic "
               "limit, now at parity (frame reduction closed the last regression).** "
               "The headline metric is the intra-step **transient** `max_block`. With "
               "frame reduction ON (default), `cultivation_d5`'s transient peak is `2^10` "
               "= Clifft `2^10` — **parity, no longer a loss** (the earlier `2^11` 2x-loss "
               "was the *pre-reduction* number, now removed). The settled **resident** "
               "dips to `2^9` between measurements (a 2x *sub-peak* factorisation win: the "
               "10 active idents no longer fit one block once measured-out qubits are "
               "decoupled), but the memory-provisioning peak is the transient `2^10`, so "
               "the honest headline is **parity, not a win** — the magic is irreducible. "
               "Block factoring guarantees the settled `max_block` never *exceeds* Clifft "
               "on any circuit; the only residual is a per-measurement, **sub-peak** "
               "transient `+1` at the lone measurement where Clifft's local rank dips "
               "(`cultivation_d5` meas 3: Clifft k=2 -> NC 3) — see the "
               "measurement-dependency report; it never reaches the global peak, so it "
               "does not change feasibility. The other `<1x` cell, `cultivation_d3` MEMORY "
               "`0.62x`, is a polynomial-overhead asymmetry (active dimension `2^4 = 2^4` "
               "parity; the byte baseline counts NC's tableau+pending but not Clifft's own "
               "stabilizer overhead; the exponential term does not lose).\n")
    out.append("\n## Transient & resident peak `max_block` vs Clifft "
               "(the honest no-regression picture)\n")
    out.append("> `transient` is the headline intra-step high-water (memory-provisioning "
               "peak); `resident` is the settled step-boundary block. With frame reduction "
               "ON, **no circuit's transient peak exceeds Clifft** — `cultivation_d5` is "
               "parity (`2^10 = 2^10`) and every other circuit is a win; the settled "
               "resident is parity-or-win everywhere (`cultivation_d5` settles to `2^9`).\n")
    out.append(md_noreg(noreg_rows()))
    out.append("\n## ACTIVE-STATE SIZE  (dense-equivalent dimension, 2^x — NC = intra-step TRANSIENT peak)\n")
    out.append(md_table(metric_rows(**active)))
    out.append("\n## MEMORY  (bytes — NC includes tableau+pending overhead; Clifft = dense `16·2^k` only)\n")
    out.append(md_table(metric_rows(**memory)))
    text = "\n".join(out) + "\n"
    open(os.path.join(SRC, "DETAILED_TABLE.md"), "w").write(text)
    return text


def write_xlsx(active, memory):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="D9E1F2")
    adv_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    def sheet(ws, title, rows):
        ws.title = title
        # group header row: circuit | PEAK (5) | SUM (5)
        ws.append(["", "PEAK", "", "", "", "", "SUM", "", "", "", ""])
        ws.merge_cells("B1:F1")
        ws.merge_cells("G1:K1")
        ws.append(HEADERS)
        for r in rows:
            ws.append(r)
        # style group header
        for col in (2, 7):
            cell = ws.cell(row=1, column=col)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = center
        # style column header
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = sub_fill
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
        # style body
        adv_cols = {4, 6, 9, 11}
        for ri in range(3, 3 + len(rows)):
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=ri, column=col)
                cell.border = border
                cell.alignment = right if col > 1 else Alignment(horizontal="left")
                if col in adv_cols:
                    cell.fill = adv_fill
                    cell.font = Font(bold=True)
        # widths
        ws.column_dimensions["A"].width = 18
        for col in range(2, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "B3"

    def simple_sheet(ws, title, headers, rows):
        ws.title = title
        ws.append(headers)
        for r in rows:
            ws.append(r)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = sub_fill
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
        for ri in range(2, 2 + len(rows)):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ri, column=col)
                cell.border = border
                cell.alignment = right if col > 1 else Alignment(horizontal="left")
                v = str(cell.value)
                if "LOSS" in v or "win" in v or v == "parity":   # verdict cells
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor=(
                        "FFC7CE" if "LOSS" in v else
                        "FFF2CC" if v == "parity" else "E2EFDA"))
        ws.column_dimensions["A"].width = 18
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.freeze_panes = "B2"

    simple_sheet(wb.active, "Transient vs Resident", NOREG_HEADERS, noreg_rows())
    sheet(wb.create_sheet(), "Active-State (2^x)", metric_rows(**active))
    sheet(wb.create_sheet(), "Memory (bytes)", metric_rows(**memory))
    path = os.path.join(SRC, "DETAILED_TABLE.xlsx")
    wb.save(path)
    return path


def main():
    text = write_md(ACTIVE, MEMORY)
    path = write_xlsx(ACTIVE, MEMORY)
    print(text)
    print(f"\nExcel written: {path}")


if __name__ == "__main__":
    main()
